'''NOTES
'''

# creating exexutable: pyinstaller.exe --onefile --noconsole InvoiceHelper_v3.py

# pathing
import os
from os import path
import shutil
import sys
# tkinter
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
# data
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
# automation
import pyautogui as auto
import time
# from PIL import Image 
# misc
import itertools
import clipboard
import re
from ast import literal_eval
import math

### IS
#%%
def select_folder1():
    global curr_folder
    global curr_csv
    global curr_xls
    folder = filedialog.askdirectory(title='Select Reading Folder',initialdir='invoices')
    curr_folder = folder
    curr_csv = os.path.join(curr_folder,"invoice_info.csv")
    curr_xls = os.path.join(curr_folder,"invoice_info.xlsx")

    display_folder = folder.split("/")[-1]
    choice1.set("Selected: " + display_folder)
    # print(display_folder)

    display.config(state=NORMAL)
    s = '\nReading Folder selected: ' + display_folder + '\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def clean_names():
    global curr_folder 
    copy = 0
 
    display.config(state=NORMAL)
    s = 'Cleaning Names...\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)
    
    remove_list = display2.get(1.0, END).replace('phrases to remove:\n','').replace('\n','').split(',')
    
    # rename files in folder
    for filename in os.listdir(curr_folder): 
        noext = filename.upper()
        noext = noext.split('.')
        noext.pop() # drop extension
        noext = ''.join(noext)
        
        src = curr_folder + '\\' + filename.upper() 
        dst = noext
        # print(dst)
        for r in remove_list:
            # print(r)
            dst = dst.replace(r.upper(),'')
        # if cleaned file name alr exists and not equal to source add copy #
        if dst + '.pdf' in os.listdir(curr_folder) and dst + '.pdf' != filename:
            dst = dst + 'COPY' + str(copy)
            copy += 1
        
        dst = curr_folder + '\\' + dst + '.pdf'
        os.rename(src, dst)
    # next create csv
    display.config(state=NORMAL)
    s = '\nNames cleaned! Verify with Open Folder\nThen click Create CSV\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)
    
def create_csv():
    '''
    get invoice names from temp -> insert into df -> to csv
    '''
    global curr_folder
    global file_list
    global df
    global curr_csv
    # remove pdf extension
    file_list = os.listdir(curr_folder) # list of all files in temp folder
    file_list = [f.upper().split('.')[0] for f in file_list if f.split('.')[1] == 'pdf'] # only add pdf
    
    # display message
    display.config(state=NORMAL)
    s = 'Creating CSV...\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)
    
    df = pd.DataFrame()
    # insert invoice numbers as first column and initialize csv
    df.insert(0,'invoice number',file_list) 
    
    curr_csv = os.path.join(curr_folder,'invoice_info.csv')
    df.to_csv(curr_csv, index=False)
    path2 = os.path.join(curr_folder,'invoice_info_backup.csv')
    df.to_csv(path2, index=False)
    
    # confirm that correct invoice numbers
    display.config(state=NORMAL)
    s = '\nCSV Created! Open CSV to check\nThen, Run Voyager\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def get_vouchers():
    global df
    global curr_folder
    global curr_csv
    display.config(state=NORMAL)

    # curr_csv = os.path.join(curr_folder,'invoice_info.csv')
    try:
        df = pd.read_csv(curr_csv)
    except FileNotFoundError as err:
        s = str(err) + "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    # print(df)
    # search all invoices from csv
    voucher_ids = []
    for i in df['invoice number']:        
        
        # clear search bar
        try:
            x,y = auto.locateCenterOnScreen('ref_images/voy_search.png',grayscale=True)
            auto.moveTo(x+200,y+3)
            auto.dragTo(x,y+3,duration=1)
            time.sleep(0.4)
            auto.typewrite(['backspace'])
            # enter invoice number
            auto.typewrite(str(i))
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/voy_search.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

        # search
        try:
            x,y = auto.locateCenterOnScreen('ref_images/voy_search_btn.png',grayscale=True)
            auto.click(x,y)
            time.sleep(1) # give time for closing error
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/voy_search_btn.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        
        # open invoice
        try:
            x,y = auto.locateCenterOnScreen('ref_images/edit_inv_btn.png',grayscale=True)
            auto.click(x,y)
            time.sleep(1)
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/edit_inv_btn.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

        # highlight voucher id and copy
        try:
            x,y = auto.locateCenterOnScreen('ref_images/voy_vid.png',grayscale=True)
            auto.moveTo(x+100,y)
            auto.dragTo(x,y,duration=0.6)
            auto.hotkey('ctrl','c')
            # store voucher id from clipboard
            time.sleep(0.2)
            v_id = clipboard.paste()
            voucher_ids.append(v_id)
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/voy_vid.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

        # close window
        try:
            x,y = auto.locateCenterOnScreen('ref_images/voy_close.png',grayscale=True)
            # x,y = 1253,357
            auto.click(x,y)
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/voy_close.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

    # # insert voucher_id column into df and update csv
    df.insert(1,'voucher id', voucher_ids)
    df.to_csv(curr_csv, index=False)
    backup = os.path.join(curr_folder,'invoice_info_backup.csv')
    df.to_csv(backup, index=False)

    
    s = '\nVoucher IDs obtained:\n'
    display.insert(END, s, ('b'))
    s = ''
    for v in voucher_ids:
        s = s + v + '\n'
    s = s + 'Add to invoice names with Add Voucher IDs\nThen, Run Jasper\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)
    
def add_vids():
    global curr_folder
    global df
    global curr_csv
    global file_list
    display.config(state=NORMAL)

    try:
        df = pd.read_csv(curr_csv)
    except FileNotFoundError as err:
        s = "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return

    for idx, row in df.iterrows():
        # print(row)
        v = row['voucher id']
        inv = row['invoice number']
        # print(v,inv)
        f_old = os.path.join(curr_folder,str(inv) + '.pdf')
        f_new = os.path.join(curr_folder,str(v)+'_'+str(inv) + '.pdf')
        # only rename if not already renamed
        if not os.path.exists(f_new):
            try:
                os.rename(f_old,f_new) # prepend v_id 
            except FileNotFoundError as err:
                s = "\nDidn't find the invoice: " + f_old + "\n"
                display.insert(END, s, ('e'))

    s = '\nVoucher IDs added! Verify with Open Folder\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def run_jasper():
    global df
    global curr_folder
    global curr_csv
    display.config(state=NORMAL)

    try:
        df = pd.read_csv(curr_csv)
    except FileNotFoundError as err:
        s = "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return

    # search for Invoice search
    try:
        x,y = auto.locateCenterOnScreen('ref_images/jasp_search_ref.png',grayscale=True)
        auto.click(x+30,y-3)
        auto.typewrite('Voyager Invoice Search')
        auto.typewrite(['enter'])
        time.sleep(3) # wait for load
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_search_ref.png\n'
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    # open
    try:
        x,y = auto.locateCenterOnScreen('ref_images/jasp_VIS.png',grayscale=True)
        # x,y = 471,236
        auto.moveTo(x,y)
        # time.sleep(2) # just in case
        auto.click(x,y) # open VIS
        time.sleep(5) # wait for load
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_VIS.png\n'
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    # close first search window
    try:
        x,y = auto.locateCenterOnScreen('ref_images/jasp_cancel.png',grayscale=True)
        # x,y = 990,968
        auto.click(x,y)
        time.sleep(1)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_cancel.png\n'
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    # loop through all voucher ID's to get info
    dates = []
    inv_nums = []
    vouch_ids = []
    tax_codes = []
    faus = [] # can use to verify tax code
    fund_codes = []
    amounts = []
    totals = []
    vendor_taxes = []
    state_taxes = []
    for v in df['voucher id']:
        # open search window
        try:
            x,y = auto.locateCenterOnScreen('ref_images/jasp_search_window.png',grayscale=True)
            auto.click(x,y)     
        except:
            try:
                x,y = auto.locateCenterOnScreen('ref_images/jasp_search_window2.png',grayscale=True)
                auto.click(x,y)
            except TypeError as err:
                s = str(err) + '\nPyAutoGui could not find either:\n ref_images/jasp_search_window2.png\n or ref_images/jasp_search_window2.png\n for voucher ' + str(v)
                display.insert(END, s, ('e'))
                display.config(state=DISABLED)
                return
        time.sleep(1)

        # clear last saerch
        try:
            x,y = auto.locateCenterOnScreen('ref_images/jasp_search_field.png',grayscale=True)
            auto.moveTo(x+100,y+30)
            auto.dragTo(x-70,y+30,duration=0.8)
            time.sleep(0.4)
            auto.typewrite(['backspace'])    
            # insert new search
            auto.typewrite(str(v))
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_search_field.png\n for voucher ' + str(v)
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        
        # click ok
        try:
            x,y = auto.locateCenterOnScreen('ref_images/jasp_ok.png',grayscale=True)
            auto.click(x,y)
            time.sleep(2) # wait for loading to show
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_ok.png\n for voucher ' + str(v)
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

        try:
            counter = 0
            while auto.locateOnScreen('ref_images/jasp_load.png',grayscale=True) and counter <= 30:
                time.sleep(1)
                counter += 1
            if counter > 30:
                print("timed out waiting for jasper")
                return
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_load.png\n for voucher ' + str(v)
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        
        # first delete previous jasp csv file if exists
        dl_path_file = 'ref_txts/download_path.txt'
        with open(dl_path_file) as f:
            jasp_csv = f.readline().strip('\n')
        if path.exists(jasp_csv):
            os.remove(jasp_csv)

        # save as csv to downloads
        try:
            x,y = auto.locateCenterOnScreen('ref_images/jasp_export.png',grayscale=True)
            auto.moveTo(x,y) # hover over export
            time.sleep(1)
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/jasp_export.png\n for voucher ' + str(v)
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

        try:
            x,y = auto.locateCenterOnScreen('ref_images/jasp_csv.png',grayscale=True)
            auto.click(x,y) # select csv
        except:
            dates.append('')
            inv_nums.append('')
            vouch_ids.append(str(v))
            tax_codes.append('')
            faus.append('')
            fund_codes.append('')
            amounts.append('')
            totals.append('')
            vendor_taxes.append('')
            state_taxes.append('')
            df2 = pd.DataFrame() # reinit df
            df2.insert(0,'invoice date', dates)
            df2.insert(1,'invoice num', inv_nums)
            df2.insert(2,'voucher id', vouch_ids)
            df2.insert(3,'FAU', faus)
            df2.insert(4,'fund code', fund_codes)
            df2.insert(5,'amount', amounts)    
            df2.insert(6,'tax code', tax_codes)
            df2.insert(7,'vendor tax', vendor_taxes)
            df2.insert(8,'state tax', state_taxes)
            df2.insert(9,'total', totals)
            backup = os.path.join(curr_folder,'invoice_info_backup.csv')
            df2.to_csv(backup,index=False)

            s = str(err) + '\nPyAutoGui did not find: ref_images/jasp_csv.png\n for voucher ' + str(v)
            display.insert(END, s, ('e'))

            continue
        time.sleep(2) # wait for download

        # x,y = auto.locateCenterOnScreen('ref_images/jasp_close_dl_ref.png',grayscale=True)
        # auto.click(x,y+10) # close popup (popup displaces coordinates)

        # obtain info from jasper csv
        voy_df = pd.read_csv(jasp_csv)
        dates.extend(voy_df['INVOICE DATE'].tolist())
        inv_nums.extend(voy_df['INVOICE NUMBER'].tolist())
        vouch_ids.extend(voy_df['INVOICE ID'].tolist())
        tax_codes.extend(voy_df['TAX CODE'].tolist())
        faus.extend(voy_df['FAU'].tolist())
        fund_codes.extend(voy_df['FUND CODE'].tolist())
        amounts.extend(voy_df['AMOUNT'].tolist())
        totals.extend(voy_df['TOTAL'].tolist())
        vendor_taxes.extend(voy_df['VENDOR TAX'].tolist())
        state_taxes.extend(voy_df['STATE TAX'].tolist())

        df2 = pd.DataFrame() # reinit df
        df2.insert(0,'invoice date', dates)
        df2.insert(1,'invoice num', inv_nums)
        df2.insert(2,'voucher id', vouch_ids)
        df2.insert(3,'FAU', faus)
        df2.insert(4,'fund code', fund_codes)
        df2.insert(5,'amount', amounts)    
        df2.insert(6,'tax code', tax_codes)
        df2.insert(7,'vendor tax', vendor_taxes)
        df2.insert(8,'state tax', state_taxes)
        df2.insert(9,'total', totals)
        backup = os.path.join(curr_folder,'invoice_info_backup.csv')
        df2.to_csv(backup,index=False)
          
    # print('here2')
    # add new columns into df
    df = pd.DataFrame() # reinit
    df.insert(0,'invoice date', dates)
    df.insert(1,'invoice num', inv_nums)
    df.insert(2,'voucher id', vouch_ids)
    df.insert(3,'FAU', faus)
    df.insert(4,'fund code', fund_codes)
    df.insert(5,'amount', amounts)    
    df.insert(6,'tax code', tax_codes)
    df.insert(7,'vendor tax', vendor_taxes)
    df.insert(8,'state tax', state_taxes)
    df.insert(9,'total', totals)
    
    # make easier for notes, pre-add colon
    df['fund code'] = [str(f) + ':' for f in df['fund code']]
    df['tax code'] = [str(t) + ':' if t != 'EX' else t for t in df['tax code']]
    # confirm final csv
    
    s = '\nFinished grabbing info from Jasper!\n'
    display.insert(END, s, ('b'))
    
    df.to_csv(curr_csv,index=False)

    s = '\nCSV created!\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)
    
def create_excel():
    global df
    global curr_folder
    global curr_csv
    global curr_xls
    display.config(state=NORMAL)

    try:
        df = pd.read_csv(curr_csv)
    except FileNotFoundError:
        s = "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    df = df.rename(columns={'tax code': 'charge type','vendor tax': 'tax 1', 'state tax': 'tax 2'})
    df['tax code'] = df['charge type']
    df['tax code'].replace(['EX','VR:'], '',inplace=True)
    cols = df.columns.tolist()
    # print(cols)
    cols = cols[:7] + cols[-1:] + cols[7:-1]
    df = df[cols]
    # print(cols)
    # df.insert(6,'tax code')

    voucher_list = np.unique(df['voucher id'])
    curr_xls = os.path.join(curr_folder,'invoice_info.xlsx')
    # df.to_excel(curr_xls,index=False) # create initial xlsx
    
    # create xls with one sheet per invoice
    writer = pd.ExcelWriter(curr_xls)
    for vouch in voucher_list:
        df_sub = df.loc[df['voucher id']==vouch]
        df_sub.to_excel(writer,str(vouch),index=False)
    writer.save()

    # adjust column widths to fit content and color
    wb = openpyxl.load_workbook(filename = curr_xls)     
    for ws in wb.worksheets:   
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 3
        for col in ws.iter_cols(min_col=8, max_col=10, min_row=2, max_row=None):
            for cell in col:
                cell.fill = PatternFill(fgColor="FFC7CE", fill_type = "mediumGray")
    wb.save(curr_xls) # update excel

    # backup only if DNE
    backup = os.path.join(curr_folder,"invoice_info_backup.xlsx")
    if not os.path.exists(backup):
        df = pd.read_excel(curr_xls)
        df.to_excel(backup)

    
    s = '\nExcel file created!\n'
    display.insert(END, s, ('b'))
    s = 'Open XLS to check\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def add_notes():
    global df
    global file_list
    global curr_xls
    global curr_folder
    display.config(state=NORMAL)

    try:
        df = pd.read_csv(curr_csv)
    except FileNotFoundError:
        s = "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    # open invoices in reading folder
    file_list = os.listdir(curr_folder)
    # filter out csvs, excels
    file_list = [f for f in file_list if not f.startswith('i')]
    v_ids = [f.split('_')[0] for f in file_list]
    unique = []
    duplicate = []
    for v in v_ids:
        if v in unique:
            duplicate.append(v)
        unique.append(v)
    # filter out already added files:
    file_list = [f for f in file_list if f.split('_')[0] not in duplicate] 

    for f in file_list:
    # f = file_list[0]
        note_items = '' # holds notes for file
        v_id = f.split('_')[0]
        # prepare notes from sheet
        try:
            v_info = pd.read_excel(curr_xls,sheet_name=v_id)
        except FileNotFoundError:
            s = "\nCan't find invoice_info.xlsx in the folder: " + curr_folder + "\n"
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        for i,row in v_info.iterrows():
            if i != 0:
                note_items += '\n'
            note_items += row['fund code'] + ' '
            note_items += row['amount'] + '\n'
            note_items += row['charge type']
            if row['charge type'] != 'EX':
                t1 = row['tax 1']
                t2 = row['tax 2']
                t1 = '' if isinstance(t1,float) else row['tax 1']
                t2 = '' if isinstance(t2,float) else row['tax 2']
                note_items += str(t1) + str(t2) + '\n'
        # print(note_items)
        path = os.path.join(curr_folder,f)
        os.startfile(path)
        time.sleep(2)
        auto.hotkey('alt','s','f') # switch to sign mode
        time.sleep(1)
        try:
            x,y = auto.locateCenterOnScreen('ref_images/add_notes_ref_1.png',grayscale=True)
            # add v_id
            auto.click(x-160,y+40)
            auto.typewrite(v_id)
             # scroll to last page
            auto.moveTo(x-2,y+50)
            auto.dragTo(x-2,y+1000,duration=2)
            # add notes
            auto.doubleClick(x-400,y+200)  
            auto.typewrite(note_items) 
            # save invoice
            time.sleep(1)
            auto.hotkey('ctrl','s')
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/add_notes_ref_1.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        # wait for save window
        try:
            counter = 0
            while auto.locateOnScreen('ref_images/temp_save.png',grayscale=True) == None and counter <= 10:
                time.sleep(1)
                counter += 1
            # save in selected folder (first)
            auto.typewrite(['enter'])
            time.sleep(1)
            # reposition to before .pdf
            auto.press('right')
            auto.keyDown('ctrl')
            auto.press('left')
            auto.keyUp('ctrl')
            auto.press('left')
            # add initials
            try:
                init_file = 'ref_txts/initials.txt'
                with open(init_file) as f:
                    initials = f.readline().strip('\n')
                auto.typewrite('_R_' + initials)
            except FileNotFoundError:
                s = "\nCan't find initials.txt in the folder: " + curr_folder + "\n"
                display.insert(END, s, ('e'))
                display.config(state=DISABLED)
                return
             # save file
            auto.typewrite(['enter'])
            time.sleep(0.5) 
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/temp_save.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return
        # close pdf
        try:
            x,y = auto.locateCenterOnScreen('ref_images/add_notes_close.png',grayscale=True)
            auto.click(x,y)
        except TypeError as err:
            s = str(err) + '\nPyAutoGui could not find: ref_images/add_notes_close.png\n'
            display.insert(END, s, ('e'))
            display.config(state=DISABLED)
            return

    # confirm final csv
    s = '\nNotes added to invoices!\n'
    display.insert(END, s, ('b'))
    s = 'Verify using Open Folder\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def open_folder():
    global curr_folder
    display.config(state=NORMAL)
    
    s = '\nOpening reading folder...\n'
    display.insert(END, s, ('a'))
    try:
        os.startfile(curr_folder)
    except FileNotFoundError as err:
        s = str(err) + "\nCan't find the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return
    
    s = 'Folder opened!\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def open_csv():
    global curr_folder
    global curr_csv
    display.config(state=NORMAL)

    s = 'Opening CSV...\n'
    display.insert(END, s, ('a'))
    try:
        os.startfile(curr_csv)
    except FileNotFoundError as err:
        s = str(err) + "\nCan't find invoice_info.csv in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return

    s = 'CSV file opened!\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def open_excel():
    global curr_xls
    display.config(state=NORMAL)

    s = 'Opening excel...\n'
    display.insert(END, s, ('a'))
    try:
        os.startfile(curr_xls)
    except FileNotFoundError as err:
        s = str(err) + "\nCan't find invoice_info.xlsx in the folder: " + curr_folder + "\n"
        display.insert(END, s, ('e'))
        display.config(state=DISABLED)
        return

    s = 'Excel file opened!\n'
    display.insert(END, s, ('a'))
    display.config(state=DISABLED)

def destroy(x=0):
    root.destroy()

### CA
#%%
def get_spec_code(spec_code,typ):
    if typ == 'Freight':
        spec_code = spec_code[:2] + 'X' + spec_code[3:]
        spec_code = spec_code[:8] + 'F8'
        spec_code = spec_code.replace('\n','')
    elif typ == 'Mono Tax':
        spec_code = spec_code[:2] + 'M' + spec_code[3:]
        spec_code = spec_code[:8] + 'T9'
        spec_code = spec_code.replace('\n','')
    elif typ == 'Serial Tax':
        spec_code = spec_code[:2] + 'S' + spec_code[3:]
        spec_code = spec_code[:8] + 'T9'
        spec_code = spec_code.replace('\n','')
    return spec_code

def init_CB():
    global vouchers
    global inv_info
    global curr_folder
    global curr_xls

    print(curr_xls)
    wb = openpyxl.load_workbook(filename = curr_xls) 
    vouchers = wb.sheetnames
    voucher_CB.delete(0, 'end')
    voucher_CB.set_completion_list(vouchers)
    voucher_CB.insert(0, vouchers[0])
    voucher_CB.focus_set()

    nc_display.config(state=NORMAL)
    s = '\nInvoices loaded into combobox!\n'
    nc_display.insert(END, s, ('a'))
    nc_display.config(state=DISABLED)

def select_folder2():
    global curr_folder
    global curr_csv
    global curr_xls
    folder = filedialog.askdirectory(title='Select Reading Folder',initialdir='invoices')
    curr_folder = folder
    curr_csv = os.path.join(folder,'invoice_info.csv')
    curr_xls = os.path.join(folder,'invoice_info.xlsx')
    display_folder = folder.split("/")[-1]
    
    choice2.set("Selected: " + display_folder)
    # print(display_folder)

    nc_display.config(state=NORMAL)
    s = '\nReading Folder selected: ' + display_folder + '\n'
    nc_display.insert(END, s, ('a'))
    nc_display.config(state=DISABLED)

    init_CB()

def open_inv():
    global voucher_CB
    v_id = voucher_CB.get()
    nc_display.config(state=NORMAL)

    # reset charge values
    charge_num.delete("1.0","end")
    charge_num.insert(END, (2))
    charge_inc.delete("1.0","end")
    charge_inc.insert(END, (1))
    library_CB.delete(library_CB.position, END)

    # search voucher in voyager and open
    # click search bar (clear first)
    try:
        x,y = auto.locateCenterOnScreen('ref_images/voy_search.png',grayscale=True)
        auto.moveTo(x+200,y+3)
        auto.dragTo(x,y+3,duration=1)
        time.sleep(0.4)
        auto.typewrite(['backspace'])
        # enter voucher number
        auto.typewrite(v_id)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/voy_search.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    # search btn
    try:
        x,y = auto.locateCenterOnScreen('ref_images/voy_search_btn.png',grayscale=True)
        auto.click(x,y)
        time.sleep(1) # give time for closing error
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/voy_search_btn.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return
        
    # open invoice
    try:
        x,y = auto.locateCenterOnScreen('ref_images/open_inv.png',grayscale=True) # USE FOR DIST
        # x,y = 1312,832
        auto.click(x,y)
        time.sleep(1)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/open_inv.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    # open other charges
    try:
        x,y = auto.locateCenterOnScreen('ref_images/other_charges1.png' ,grayscale=True) # dots
        auto.click(x,y)
        time.sleep(0.5)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/other_charges1.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return
    try:
        x,y = auto.locateCenterOnScreen('ref_images/other_charges2.png',grayscale=True) # adjust
        auto.click(x,y)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/other_charges2.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    s = '\nInvoice {} opened\n'.format(v_id)
    nc_display.insert(END, s, ('a'))
    nc_display.config(state=DISABLED)

def add_next_charge():
    global cc_df
    global fc_df
    global curr_xls
    global voucher_CB
    nc_display.config(state=NORMAL)

    v_id = voucher_CB.get()
    # get sheet specified by voucher id
    sheet_df = pd.read_excel(curr_xls,sheet_name=v_id) 
    # specify row
    row = int(charge_num.get('1.0','end-1c'))
    charge_row = sheet_df.iloc[[row-2]] # - 1 for headers, -1 for 0 indexing in df
    # print(charge_row)

    # determine charge type
    charge = str(charge_row.iloc[0]['charge type'])
    if charge[-1] != ':':
        charge = charge + ':'
    charge_typewrite = cc_df.loc[cc_df['charge type'] == charge,'typewrite'].values[0]

    # determine charge amt
    charge_amt = str(charge_row.iloc[0]['tax 1'])
    charge_amt2 = str(charge_row.iloc[0]['tax 2'])
    if charge_amt == 'nan': # get non empty charge
        charge_amt = charge_amt2
    charge_amt = str(charge_amt)

    # using library and tax code, determine fund code for tax
    library = library_CB.get()
    library_typewrite = cc_df.loc[cc_df['library'] == library,'l_typewrite'].values[0]
    tax_code = str(charge_row.iloc[0]['tax code'])
    tax_dict = {'TM:':'Mono Tax','TS':'Serial Tax','FRT':'Freight'}
    tax_code = tax_dict[tax_code]
    if charge in ['PROC_CHARGE:','EX_PROC_CHARGE:']:
        fund_code = clipboard.paste() # get code from copy
    if library in ('Contracts & Grants', 'Sales & Services', 'Special Foundation','Special Regental'):
        fund_code = clipboard.paste() # get code from copy
        # then modify for tax code
        fund_code = get_spec_code(fund_code,tax_code)
    else:
        fund_code = str(fc_df[fc_df['Library'] == library][tax_code].values[0])
    
    # click add
    try:
        x,y = auto.locateCenterOnScreen('ref_images/other_charges3.png',grayscale=True)
        auto.click(x,y)
        time.sleep(4)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/other_charges3.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return
    try:
        chevs = list(auto.locateAllOnScreen('ref_images/add_charge_chevron.png',grayscale=True)) # 4 total, first is account
        print(chevs)
        ch_type = chevs[1]
        app_mthd = chevs[2]
        ledger = chevs[3]
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/add_charge_chevron.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    # select charge type
    x,y = (ch_type[0] + ch_type[2]/2 ,ch_type[1] + ch_type[3]/2)
    auto.click(x,y)
    auto.typewrite(charge_typewrite)
    auto.click(x,y)

    # application method -> select amount
    x,y = (app_mthd[0] + app_mthd[2]/2 ,app_mthd[1] + app_mthd[3]/2)
    auto.click(x,y)
    auto.typewrite('a')
    auto.click(x,y)

    # insert amount
    try:
        x,y = auto.locateCenterOnScreen('ref_images/add_charge_amount.png',grayscale=True)
        auto.click(x,y)
        auto.typewrite(charge_amt)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/add_charge_amount.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    # select ledger
    x,y = (ledger[0] + ledger[2]/2 ,ledger[1] + ledger[3]/2)
    auto.click(x,y)
    auto.typewrite(library_typewrite)
    auto.click(x,y)

    # insert fund code
    try:
        x,y = auto.locateCenterOnScreen('ref_images/add_charge_commit_to_fund.png',grayscale=True)
        auto.click(x,y)
        auto.typewrite(fund_code)
    except TypeError as err:
        s = str(err) + '\nPyAutoGui could not find: ref_images/add_charge_commit_to_fund.png\n'
        nc_display.insert(END, s, ('e'))
        nc_display.config(state=DISABLED)
        return

    # increment row count and add to text box
    inc = int(charge_inc.get('1.0','end-1c')) # get amount to increment
    row += inc # increment
    charge_num.delete("1.0","end")
    charge_num.insert(END, str(row))

    nc_display.config(state=DISABLED)

### TCF
#%%
tkinter_umlauts=['odiaeresis', 'adiaeresis', 'udiaeresis', 'Odiaeresis', 'Adiaeresis', 'Udiaeresis', 'ssharp']

class AutocompleteEntry(Entry):
        """
        Subclass of Entry that features autocompletion.

        To enable autocompletion use set_completion_list(list) to define
        a list of possible strings to hit.
        To cycle through hits use down and up arrow keys.
        """
        def set_completion_list(self, completion_list):
                self._completion_list = sorted(completion_list, key=str.lower) # Work with a sorted list
                self._hits = []
                self._hit_index = 0
                self.position = 0
                self.bind('<KeyRelease>', self.handle_keyrelease)

        def autocomplete(self, delta=0):
                """autocomplete the Entry, delta may be 0/1/-1 to cycle through possible hits"""
                if delta: # need to delete selection otherwise we would fix the current position
                        self.delete(self.position, END)
                else: # set position to end so selection starts where textentry ended
                        self.position = len(self.get())
                # collect hits
                _hits = []
                for element in self._completion_list:
                        if element.lower().startswith(self.get().lower()):  # Match case-insensitively
                                _hits.append(element)
                # if we have a new hit list, keep this in mind
                if _hits != self._hits:
                        self._hit_index = 0
                        self._hits=_hits
                # only allow cycling if we are in a known hit list
                if _hits == self._hits and self._hits:
                        self._hit_index = (self._hit_index + delta) % len(self._hits)
                # now finally perform the auto completion
                if self._hits:
                        self.delete(0,END)
                        self.insert(0,self._hits[self._hit_index])
                        self.select_range(self.position,END)

        def handle_keyrelease(self, event):
                """event handler for the keyrelease event on this widget"""
                if event.keysym == "BackSpace":
                        self.delete(0, END)
                        # self.position = self.index(END)
                if event.keysym == "Left":
                        if self.position < self.index(END): # delete the selection
                                self.delete(self.position, END)
                        else:
                                self.position = self.position-1 # delete one character
                                self.delete(self.position, END)
                if event.keysym == "Right":
                        self.position = self.index(END) # go to end (no selection)
                if event.keysym == "Down":
                        self.autocomplete(1) # cycle to next hit
                if event.keysym == "Up":
                        self.autocomplete(-1) # cycle to previous hit
                if len(event.keysym) == 1 or event.keysym in tkinter_umlauts:
                        self.autocomplete()

class AutocompleteCombobox(ttk.Combobox):

        def set_completion_list(self, completion_list):
                """Use our completion list as our drop down selection menu, arrows move through menu."""
                self._completion_list = sorted(completion_list, key=str.lower) # Work with a sorted list
                self._hits = []
                self._hit_index = 0
                self.position = 0
                self.bind('<KeyRelease>', self.handle_keyrelease)
                self['values'] = self._completion_list  # Setup our popup menu

        def autocomplete(self, delta=0):
                """autocomplete the Combobox, delta may be 0/1/-1 to cycle through possible hits"""
                if delta: # need to delete selection otherwise we would fix the current position
                        self.delete(self.position, END)
                else: # set position to end so selection starts where textentry ended
                        self.position = len(self.get())
                # collect hits
                _hits = []
                for element in self._completion_list:
                        if element.lower().startswith(self.get().lower()): # Match case insensitively
                                _hits.append(element)
                # if we have a new hit list, keep this in mind
                if _hits != self._hits:
                        self._hit_index = 0
                        self._hits=_hits
                # only allow cycling if we are in a known hit list
                if _hits == self._hits and self._hits:
                        self._hit_index = (self._hit_index + delta) % len(self._hits)
                # now finally perform the auto completion
                if self._hits:
                        self.delete(0,END)
                        self.insert(0,self._hits[self._hit_index])
                        self.select_range(self.position,END)

        def handle_keyrelease(self, event):
                """event handler for the keyrelease event on this widget"""
                if event.keysym == "BackSpace":
                        self.delete(self.index(INSERT), END)
                        self.position = self.index(END)
                if event.keysym == "Left":
                        if self.position < self.index(END): # delete the selection
                                self.delete(self.position, END)
                        else:
                                self.position = self.position-1 # delete one character
                                self.delete(self.position, END)
                if event.keysym == "Right":
                        self.position = self.index(END) # go to end (no selection)
                if len(event.keysym) == 1:
                        self.autocomplete()

def update_fc(x):
    global spec_vis
    global fc_df
    lib = fund_CB.get()
    typ = type_CB.get()
    code = str(fc_df[fc_df['Library'] == lib][typ])[2:].replace(' ','')[:10]
    spec_code = spec_display.get("1.0","end")       
    
    if lib in ('Contracts & Grants', 'Sales & Services', 'Special Foundation','Special Regental'):
        if not spec_vis:
            spec_display.grid(row=2,column=0,padx=(0,0),pady=10)
            spec_display.config(state=NORMAL)
            s = 'Paste Spec Code'
            spec_display.insert(END, s,"b")
            spec_vis = True
    else:
        spec_display.grid_forget()
        spec_vis = False
        spec_code = '^'
        fc_display.config(state=NORMAL)
        s = code
        fc_display.delete("1.0","end")
        fc_display.insert(END, s, 'a')
        fc_display.config(state=DISABLED)
    
    if spec_code[0] in ('C','S','F','R'):
        # print(spec_code)
        spec_display.delete("1.0","end")
        spec_code = get_spec_code(spec_code,typ)    
        fc_display.config(state=NORMAL)
        fc_display.delete("1.0","end")
        fc_display.insert(END, spec_code, 'a')
        fc_display.config(state=DISABLED)

def to_clipboard():
    clipboard.copy(fc_display.get("1.0","end"))

### DOCS + HELP
#%%
def open_ref_csv(csv_file):
    path = os.path.join('ref_csvs',csv_file)
    os.startfile(path)

def open_ref_txt(txt_file):
    path = os.path.join('ref_txts',txt_file)
    os.startfile(path)

### TAB FNS
#%%
def on_tab_change(event):
    tab = event.widget.tab('current')['text']
    if tab == 'CSV Creator + Note Adder':
        pass
    elif tab == 'Charge Adder':
        root.bind('<Return>', lambda event=None: open_inv())
        root.bind('<Control-a>', lambda event=None: add_next_charge())
    elif tab == 'Tax Code Lookup':
        root.bind('<Return>', update_fc)
        root.bind('<Control-c>', lambda event=None: to_clipboard())
    else:
        pass

# TKINTER
#%%
# GLOBALS AND REF CSVS #
HEIGHT = '400'
WIDTH = '600'
BG_COLS = ['#708dc4','#dea75f','#86bd86','#c98fad']

fc_df = pd.read_csv('ref_csvs/fund_codes.csv')
libs = tuple(fc_df['Library'])
types = ('Freight','Mono Tax','Serial Tax')
spec_vis = False
cc_df = pd.read_excel('ref_csvs/charge_codes.xlsx')

vouchers = " " # will hold v_ids for charge_adder combo-box
inv_info = " " # will hold path for inv_info csv
curr_folder = " " # path for current working folder
# curr_folder = r'invoices\test' # path for current working folder
curr_csv = " "
# curr_csv = os.path.join(curr_folder,'invoice_info.csv')
curr_xls = " "
df = pd.DataFrame()
file_list = []

# ROOT #
root = Tk()
root.title('Invoice Helper')
root.configure(bg=BG_COLS[0])
root.geometry(WIDTH + "x" + HEIGHT)
root.resizable(False, False)
root.bind("<Control-q>",destroy)
s = ttk.Style()
s.configure('t1.TFrame', background=BG_COLS[0])
s.configure('t2.TFrame', background=BG_COLS[1])
s.configure('t3.TFrame', background=BG_COLS[2])
s.configure('t4.TFrame', background=BG_COLS[3])

tab_parent = ttk.Notebook(root)
tab1 = ttk.Frame(tab_parent,style='t1.TFrame')
tab2 = ttk.Frame(tab_parent,style='t2.TFrame')
tab3 = ttk.Frame(tab_parent,style='t3.TFrame')
tab4 = ttk.Frame(tab_parent,style='t4.TFrame')

tab_parent.add(tab1,text='CSV Creator + Note Adder')
tab_parent.add(tab2,text='Charge Adder')
tab_parent.add(tab3,text='Tax Code Lookup')
tab_parent.add(tab4,text='Docs + Help')
tab_parent.pack(expand=1,fill='both')
tab_parent.bind('<<NotebookTabChanged>>', on_tab_change) # to define custom hotkeys for each tab

# CSV Creator + Note Adder #
#%%
# citrix_button = Button(tab1, text="Open Citrix", command=open_citrix)
# citrix_button.grid(row=0, column=0, padx=(10,0), sticky=W)
# voy_button = Button(tab1, text="Open Voyager", command=lambda: open_voyager('i'))
# voy_button.grid(row=1, column=0, padx=(10,0),sticky=W)

# Col 0
create_folder_button = Button(tab1, text="1. Select Folder", command=select_folder1)
create_folder_button.grid(row=0, column=0,padx=(10,0),sticky=W)
clean_button = Button(tab1, text="2. Clean Names", command=clean_names)
clean_button.grid(row=1, column=0,padx=10,pady=10,sticky=W)
create_csv_button = Button(tab1, text="3. Create CSV", command=create_csv)
create_csv_button.grid(row=2, column=0,padx=10,pady=10,sticky=W)

# Col 1
choice1 = StringVar()
choose_lbl1 = Label(tab1,textvariable=choice1,background=BG_COLS[0])
choose_lbl1.grid(row=0,column=1,padx=10,pady=10,sticky=W)
voucher_button = Button(tab1, text="4. Get Vouchers", command=get_vouchers)
voucher_button.grid(row=1, column=1, padx=10,pady=10,sticky=W)
add_vid_button = Button(tab1, text="5. Add Voucher IDs", command=add_vids)
add_vid_button.grid(row=2, column=1,padx=10,pady=10,sticky=W) # to add

# Col 2
jasper_button = Button(tab1, text="6. Run Jasper", command=run_jasper)
jasper_button.grid(row=0, column=2, padx=10,pady=10,sticky=W)
create_xls_button = Button(tab1, text="7. Create XLS", command=create_excel)
create_xls_button.grid(row=1, column=2,padx=10,pady=10,sticky=W)
notes_button = Button(tab1, text="8. Add Notes", command=add_notes)
notes_button.grid(row=2, column=2,padx=10,pady=10,sticky=W)

# Col 3 (Open)
open_folder_button = Button(tab1, text="Open Folder", command=open_folder)
open_folder_button.grid(row=0, column=3,padx=10,pady=10,sticky=E)
open_csv_button = Button(tab1, text="Open CSV", command=open_csv)
open_csv_button.grid(row=1, column=3, padx=10,pady=10,sticky=E)
open_ex_button = Button(tab1, text="Open XLS", command=open_excel)
open_ex_button.grid(row=2, column=3, padx=10,pady=10,sticky=E)

# display windows
display = Text(tab1, bg='#ffffff', width=45,height=12)
display.tag_config("a", font=('Calibri',13))
display.tag_config("b", font=('Calibri',13,'bold'))
display.tag_config("e", font=('Calibri',13),foreground="#fc1c03")
display.grid(row=4,column=0,columnspan=3,padx=10, sticky=W)
display.config(state=NORMAL)
s = 'Welcome to Invoice Helper!\nClick one of the buttons to begin!\n'
display.insert(END, s, ('a'))
display.config(state=DISABLED)
display2 = Text(tab1, bg='#ffffff', width=25,height=13,font=('Calibri',10))
display2.tag_config("a", font=('Calibri',10))
display2.tag_config("b", font=('Calibri',10,'bold'))
display2.grid(row=4,column=3,columnspan=2,padx=10,pady=10,sticky=W)
s='phrases to remove:\n'
filter_file = 'ref_txts/filters.txt'
with open (filter_file, 'r') as filters:
    t=filters.read().replace('\n', ',')
display2.insert(END, s, ('b'))
display2.insert(END, t, ('a'))

# Charge Adder #
#%%

# Col 0
select_folder_btn = Button(tab2, text="1. Select Folder", command=select_folder2)
select_folder_btn.grid(row=0, column=0, padx=10,pady=10,sticky=W) # col 0 and 1
labelText=StringVar()
labelText.set("VID:")
labelVID=Label(tab2, textvariable=labelText,background=BG_COLS[1])
labelVID.grid(row=1,column=0,padx=10,sticky=W)
voucher_CB = AutocompleteEntry(tab2,width = 10)
voucher_CB.grid(row=2,column=0,padx=10,pady=(0,10),sticky=W)
open_inv_btn = Button(tab2, text="2. Open Invoice", command=open_inv)
open_inv_btn.grid(row=3, column=0,padx=10,pady=10,sticky=W)

# Col 1
choice2 = StringVar()
choose_lbl2 = Label(tab2,textvariable=choice2,background=BG_COLS[1])
choose_lbl2.grid(row=0,column=1,columnspan=2,padx=10,pady=10)
labelText2=StringVar()
labelText2.set("Row:")
labelRow=Label(tab2, textvariable=labelText2,background=BG_COLS[1])
labelRow.grid(row=1,column=1,padx=10,sticky=W)
charge_num = Text(tab2, bg='#ffffff', width=5,height=1)
charge_num.insert(END, (2))
charge_num.grid(row=2,column=1,padx=10,pady=(0,10),sticky=W)

# Col 2
labelText3=StringVar()
labelText3.set("Increment:")
labelInc=Label(tab2, textvariable=labelText3,background=BG_COLS[1])
labelInc.grid(row=1,column=2,padx=10,sticky=W)
charge_inc = Text(tab2, bg='#ffffff', width=5,height=1)
charge_inc.insert(END, (1))
charge_inc.grid(row=2,column=2,padx=10,pady=(0,10),sticky=W)
add_charge_btn = Button(tab2, text="2a. Add Charge", command=add_next_charge)
add_charge_btn.grid(row=3, column=1,columnspan=2,padx=10,pady=10,sticky=W)

# Col 3
open_folder_btn2 = Button(tab2, text="Open Folder", command=open_folder)
open_folder_btn2.grid(row=0, column=3,padx=10,pady=10,sticky=W)
labelText4=StringVar()
labelText4.set("Library:")
labelLib=Label(tab2, textvariable=labelText4,background=BG_COLS[1])
labelLib.grid(row=1,column=3,padx=10,sticky=W)
library_CB = AutocompleteEntry(tab2,width = 30)
library_CB.set_completion_list(libs)
library_CB.grid(row=2,column=3,columnspan=2,padx=10,pady=(0,10),sticky=W)
library_CB.focus_set()
library_CB.insert(0, libs[0])

# Col 4
open_ex_button2 = Button(tab2, text="Open XLS", command=lambda: open_excel())
open_ex_button2.grid(row=0, column=4,padx=10,pady=10,sticky=E)

# maindisplay
nc_display = Text(tab2, bg='#ffffff', width=60,height=10)
nc_display.tag_config("a", font=('Calibri',13))
nc_display.tag_config("b", font=('Calibri',13,'bold'))
nc_display.tag_config("e", font=('Calibri',13),foreground="#fc1c03")
nc_display.grid(row=4,column=0,columnspan=5,padx=10,pady=10,sticky=W)
nc_display.config(state=NORMAL)
s = 'Select the reading folder to begin\nThen add notes and charges\n'
nc_display.insert(END, s, ('a'))
nc_display.config(state=DISABLED)

# Tax Code Lookup #
#%%
# dropdown selects
fund_CB = AutocompleteEntry(tab3,width = 30)
fund_CB.set_completion_list(libs)
fund_CB.grid(row=0,column=0,padx=10,pady=10)
fund_CB.focus_set()
fund_CB.insert(0, libs[0])
# fund_CB.bind('<Button-1>',lambda x: fund_CB.delete(0, 'end'))

type_CB = AutocompleteEntry(tab3,width = 30)
type_CB.set_completion_list(types)
type_CB.grid(row=1,column=0,padx=10,pady=10)
type_CB.focus_set()
type_CB.insert(0, types[0])
# type_CB.bind('<Button-1>',lambda x: type_CB.delete(0, 'end'))

# paste area for special codes
spec_display = Text(tab3, bg='#ffffff', width=15,height=1,font=('Calibri', 13))
spec_display.tag_config("a", font=('Calibri',13))
spec_display.tag_config("b", font=('Calibri',13),foreground='#b0b0b0')
spec_display.bind('<Button-1>', lambda x: spec_display.delete("1.0","end"))

# fund code display and copy button
fc_display = Text(tab3, bg='#ffffff', width=15,height=1,font=('Calibri', 13))
fc_display.tag_config("a", font=('Calibri',13))
fc_display.tag_config("b", font=('Calibri',13,'bold'))
fc_display.grid(row=3,column=0,padx=(0,0),pady=10)
copy_btn = Button(tab3,text='Copy',width=4,height=1,command = to_clipboard)
copy_btn.grid(row=3,column=1,padx=0,pady=10)

# Docs + Help #
#%%
# Row 0
help_button = Button(tab4, text="How To Use", command= lambda: open_ref_txt("HowToManual.docx"))
help_button.grid(row=0, column=0,padx=10,pady=10,sticky=W)
edit_init_button = Button(tab4, text="Edit Initials", command=lambda: open_ref_txt("initials.txt"))
edit_init_button.grid(row=0, column=1,padx=10,pady=10,sticky=W)
edit_filters_button = Button(tab4, text="Edit Filters", command=lambda: open_ref_txt("filters.txt"))
edit_filters_button.grid(row=0, column=2,padx=10,pady=10,sticky=W)
edit_DL_button = Button(tab4, text="Edit DL Path", command=lambda: open_ref_txt("download_path.txt"))
edit_DL_button.grid(row=0, column=3,padx=10,pady=10,sticky=W)

# Row 1
charge_code_button = Button(tab4, text="View Charge Type Codes", command= lambda: open_ref_csv("charge_codes.xlsx"))
charge_code_button.grid(row=1, column=0,columnspan=2,padx=10,pady=10,sticky=W)
tax_code_button = Button(tab4, text="View Tax Codes", command=lambda: open_ref_csv("fund_codes.csv"))
tax_code_button.grid(row=1, column=2,columnspan=2,padx=10,pady=10,sticky=W)

root.mainloop()